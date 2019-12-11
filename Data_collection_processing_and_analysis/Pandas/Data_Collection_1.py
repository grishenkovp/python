import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import xlrd

# **********************************************************************************************************************
# Example 1. Чтение данных из текстового файла формата csv. Задан полный путь до искомого файла
path = 'C:/Users/Pavel/PycharmProjects/Part_2_Pandas/dataset/data1.csv'
df = pd.read_csv(path)
print(df)

# Example 2. Чтение данных из текстового файла формата csv. Задан краткий путь до искомого файла, так как он находится
# в папке dataset проекта
path = 'dataset/data1.csv'
df = pd.read_csv(path)
print(df)

# Example 3. Чтение данных из текстового файла формата csv. Выбор выбор делителя столбцов (полей) (sep)
path = 'dataset/data1.csv'
df = pd.read_csv(path, sep=';')
print(df)

# Example 4. Чтение данных из текстового файла формата csv. Выбор делителя столбцов (полей) (sep)
path = 'dataset/data2.csv'
df = pd.read_csv(path, sep='\s+')
print(df)

# Example 4. Чтение данных из текстового файла формата csv. Выбор знака, разделяющего целые и дроб. значения (decimal)
path = 'dataset/data3.csv'
df = pd.read_csv(path, sep=';', decimal=',')
print(df)
# print(df.dtypes)
# # print(df.info())

# Example 5. Чтение данных из текстового файла формата csv. Парсинг дат (parse_dates)
path = 'dataset/data3.csv'
df = pd.read_csv(path, sep=';', decimal=',', parse_dates=['date'])
print(df)

# Example 6. Чтение данных из текстового файла формата csv. Выбор столбцов (полей) для загрузки (usecols)
path = 'dataset/data3.csv'
df = pd.read_csv(path, sep=';', decimal=',', parse_dates=['date'], usecols=['date', 'product','value'])
print(df)

# Example 7. Чтение данных из текстового файла формата csv. Выбор количества загружаемых строк  (nrows)
path = 'dataset/data3.csv'
df = pd.read_csv(path, sep=';', decimal=',', parse_dates=['date'], usecols=['date', 'product','value'], nrows=4)
print(df)

# Example 8. Чтение данных из текстового файла формата txt.
path = 'dataset/data4.txt'
df = pd.read_csv(path, sep=';', decimal=',', parse_dates=['date'])
print(df)

# Example 9. Чтение данных из текстового файла формата txt. Заполнение названий столбцов
path = 'dataset/data5.txt'
df = pd.read_csv(path, sep=';', decimal=',', names = ['date','region','product','value'])
print(df)

# Example 10. Чтение данных из текстового файла формата txt. Выбор столбца-индекса
path = 'dataset/data5.txt'
df = pd.read_csv(path, sep=';', decimal=',', names = ['date','region','product','value'], index_col=['date'])
print(df)

# Example 11. Чтение данных из текстового файла формата txt. Пропуск строк
path = 'dataset/data5.txt'
df = pd.read_csv(path, sep=';', decimal=',', names = ['date','region','product','value'], skiprows=2)
print(df)
# **********************************************************************************************************************

# Example 12. Чтение данных из файла-excel. Метод Pandas
path = 'dataset/data6.xlsx'
df = pd.read_excel(path, sheet_name='Лист1')
print(df)

# Example 13. Чтение данных из файла-excel. Библиотека OpenPyXL
wb = load_workbook(filename = 'C:/Users/Pavel/PycharmProjects/Part_2_Pandas/dataset/data6.xlsx')
ws = wb['Лист1']
for cellObj in ws['A2':'D2']:
    for cell in cellObj:
        print(cell.value)

# Example 14. Чтение данных из файла-excel. Библиотека XLRD
wb = xlrd.open_workbook('C:/Users/Pavel/PycharmProjects/Part_2_Pandas/dataset/data6.xlsx')
ws = wb.sheet_by_name('Лист1')
x = ws.cell(0, 0).value
print(x)
# **********************************************************************************************************************
